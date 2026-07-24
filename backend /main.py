"""
Codiq Engine — FastAPI Backend (multi-question projects)
Run: uvicorn main:app --reload --port 8000
"""

import os, uuid, json, time, threading, random, io
from pathlib import Path

import pandas as pd
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse

app = FastAPI(title="Codiq Engine API", version="2.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True, allow_methods=["*"], allow_headers=["*"],
)

UPLOAD_DIR = Path("uploads"); UPLOAD_DIR.mkdir(exist_ok=True)
jobs: dict = {}


# ─────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────
def detect_columns(df):
    col_map = {}
    for col in df.columns:
        low = col.lower()
        if any(k in low for k in ["respid","resp_id","response_id"]) and "RESPID" not in col_map:
            col_map["RESPID"] = col
        elif any(k in low for k in ["qid","question_id","q_id"]) and "QID" not in col_map:
            col_map["QID"] = col
        elif any(k in low for k in ["verbatim","response","text","answer"]) and "Verbatim" not in col_map:
            col_map["Verbatim"] = col
    if len(col_map) < 3:
        cols = list(df.columns)
        col_map = {"RESPID": cols[0], "QID": cols[1] if len(cols)>1 else cols[0],
                   "Verbatim": cols[2] if len(cols)>2 else cols[-1]}
    return col_map


def read_rows(path):
    df = pd.read_excel(path, dtype=str).fillna("")
    df.columns = [c.strip() for c in df.columns]
    cm = detect_columns(df)
    df = df.rename(columns={v: k for k, v in cm.items()})
    df = df[["RESPID", "QID", "Verbatim"]].copy()
    df["Verbatim"] = df["Verbatim"].str.strip()
    df = df[df["Verbatim"] != ""].reset_index(drop=True)
    question = df["QID"].mode()[0] if len(df) else ""
    rows = [{"respid": str(r["RESPID"]), "qid": str(r["QID"]), "verbatim": str(r["Verbatim"])}
            for _, r in df.iterrows()]
    return rows, question


# ─────────────────────────────────────────────────────────────
#  MOCK PIPELINE — now produces CODED all_rows
#  (Replace body with real pipeline.py calls later)
# ─────────────────────────────────────────────────────────────
def mock_pipeline(job_id, topic, min_threshold):
    job = jobs[job_id]
    stages = [
        ("Discovering themes",   0.05, 0.25, 3),
        ("Consolidating themes", 0.25, 0.40, 2),
        ("Assigning NET codes",  0.40, 0.55, 2),
        ("Generating keywords",  0.55, 0.65, 1),
        ("Coding responses",     0.65, 0.88, 4),
        ("Validating output",    0.88, 0.97, 1),
        ("Writing output",       0.97, 1.00, 1),
    ]
    for name, p0, p1, dur in stages:
        if job["status"] == "cancelled": return
        job["stage"] = name; job["status"] = "running"
        for i in range(8):
            time.sleep(dur / 8)
            job["progress"] = p0 + (p1 - p0) * ((i+1)/8)
        job["log"].append(f"✓ {name} complete")

    # Build a demo codebook
    codebook = [
        {"code":201,"net":200,"net_name":"Functional Benefits","theme":"Core Product Attribute","description":""},
        {"code":202,"net":200,"net_name":"Functional Benefits","theme":"Health / Wellness Benefit","description":""},
        {"code":301,"net":300,"net_name":"Emotional Benefits","theme":"Positive Brand Feeling","description":""},
        {"code":302,"net":300,"net_name":"Emotional Benefits","theme":"Nostalgic Association","description":""},
        {"code":401,"net":400,"net_name":"Brand & Trust","theme":"Brand Recognition","description":""},
        {"code":801,"net":800,"net_name":"Functional Concerns","theme":"Taste / Quality Concern","description":""},
        {"code":901,"net":900,"net_name":"Emotional Concerns","theme":"General Dissatisfaction","description":""},
        {"code":9997,"net":9990,"net_name":"Reserved","theme":"Nothing/None/NA","description":""},
    ]
    real_codes = [201,202,301,302,401,801,901]

    # Code every row (mock: random 1-2 codes + sentiment)
    rows = job["input_rows"]
    sentiments = ["Positive","Negative","Mixed"]
    coded_rows = []
    pos = neg = mix = other = 0
    for r in rows:
        v = r["verbatim"].strip().lower()
        if v in ("nothing","none","n/a","na","nothing.","-"):
            codes = [9997]; sent = "Mixed"; other += 1
        else:
            n = random.choice([1,1,2])
            codes = random.sample(real_codes, n)
            sent  = random.choice(sentiments)
        if sent == "Positive": pos += 1
        elif sent == "Negative": neg += 1
        else: mix += 1
        coded_rows.append({**r, "sentiment": sent, "codes": codes})

    job["all_rows"] = coded_rows
    job["codebook"] = codebook
    job["stats"] = {"total": len(rows), "themes": len(real_codes),
                    "positive": pos, "negative": neg, "mixed": mix, "other": other}
    job["status"] = "done"; job["progress"] = 1.0
    job["log"].append("✓ Pipeline complete — responses coded")


# ─────────────────────────────────────────────────────────────
#  REAL PIPELINE — uses pipeline.py (Ollama/Mistral) if available
# ─────────────────────────────────────────────────────────────
def _ollama_alive():
    try:
        import requests as _rq
        r = _rq.get("http://localhost:11434/api/tags", timeout=3)
        return r.status_code == 200
    except Exception:
        return False


def real_pipeline(job_id, topic, min_threshold):
    """Full AI pipeline using pipeline.py. Falls back to mock on any import failure."""
    job = jobs[job_id]
    try:
        import sys
        here = os.path.dirname(os.path.abspath(__file__))
        if here not in sys.path:
            sys.path.insert(0, here)
        import pipeline as pl
    except Exception as e:
        job["log"].append(f"⚠ pipeline.py not importable ({e}) — using mock pipeline")
        return mock_pipeline(job_id, topic, min_threshold)

    def log(msg):
        job["log"].append(msg)

    try:
        job["status"] = "running"

        # ── Read input ─────────────────────────────────────
        job["stage"] = "Reading input"; log("Reading input file…")
        df, question, verbatims = pl.read_input(job["input_path"])
        log(f"✓ {len(df)} responses loaded")
        job["progress"] = 0.04

        # ── Stage 1: Discover ──────────────────────────────
        job["stage"] = "Discovering themes"; log("Stage 1 · Discovering themes…")
        raw_themes = pl.discover_themes(verbatims, min_threshold)
        log(f"✓ {len(raw_themes)} raw themes discovered")
        job["progress"] = 0.25
        if job["status"] == "cancelled": return

        # ── Stage 2: Consolidate ───────────────────────────
        job["stage"] = "Consolidating themes"; log("Stage 2 · Consolidating themes…")
        consolidated = pl.consolidate_themes(raw_themes)
        log(f"✓ {len(consolidated)} themes after consolidation")
        job["progress"] = 0.38

        # ── Stage 3: NET codes ─────────────────────────────
        job["stage"] = "Assigning NET codes"; log("Stage 3 · Assigning NET codes…")
        codebook = pl.assign_net_codes(consolidated)
        log(f"✓ {len(codebook)} coded themes")
        job["progress"] = 0.50

        cb_path = job["input_path"].rsplit(".", 1)[0] + "_codebook.json"
        pl.save_codebook(codebook, cb_path)

        # ── Stage 4: Keywords ──────────────────────────────
        job["stage"] = "Generating keywords"; log("Stage 4 · Generating keyword map…")
        kw_map = pl.generate_keyword_map(codebook, topic)
        log(f"✓ Keywords generated for {len(kw_map)} themes")
        job["progress"] = 0.58
        if job["status"] == "cancelled": return

        # ── Stage 5: Assign codes ──────────────────────────
        job["stage"] = "Coding responses"; log("Stage 5 · Assigning codes to responses…")
        coded_df = pl.assign_codes(df, codebook, topic)
        log("✓ Code assignment complete")
        job["progress"] = 0.85

        # ── Stage 6: Validate ──────────────────────────────
        job["stage"] = "Validating output"; log("Stage 6 · Validating & auto-correcting…")
        coded_df, fixes = pl.validate_and_fix(coded_df, codebook, kw_map)
        log(f"✓ {fixes} corrections applied")
        codebook, coded_df, merges = pl.deduplicate_codebook(codebook, coded_df)
        if merges:
            log(f"↳ Merged {len(merges)} overlapping theme(s)")
        job["progress"] = 0.94

        # ── Build all_rows for the workspace ───────────────
        job["stage"] = "Writing output"
        code_cols = [f"Code_{i}" for i in range(1, pl.MAX_CODES + 1)]
        coded_rows = []
        pos = neg = mix = 0
        for _, r in coded_df.iterrows():
            codes = []
            for c in code_cols:
                v = r.get(c)
                if pd.notna(v):
                    try: codes.append(int(float(v)))
                    except Exception: pass
            sent = str(r.get("Sentiments", "Mixed"))
            if sent == "Positive": pos += 1
            elif sent == "Negative": neg += 1
            else: mix += 1
            coded_rows.append({
                "respid":   str(r["RESPID"]),
                "qid":      str(r["QID"]),
                "verbatim": str(r["Verbatim"]),
                "sentiment": sent,
                "codes":    codes,
            })

        RESERVED = set(pl.RESERVED.keys())
        reserved_only = sum(1 for row in coded_rows
                            if row["codes"] and all(c in RESERVED for c in row["codes"]))

        job["all_rows"] = coded_rows
        job["codebook"] = codebook
        job["stats"] = {"total": len(coded_rows), "themes": len(codebook),
                        "positive": pos, "negative": neg, "mixed": mix,
                        "other": reserved_only}
        job["status"] = "done"; job["progress"] = 1.0
        log(""); log("✓ Pipeline complete — responses coded")

    except Exception as e:
        import traceback
        job["status"] = "error"
        job["log"].append(f"✗ Error: {e}")
        job["log"].append(traceback.format_exc()[-800:])


# ─────────────────────────────────────────────────────────────
#  ROUTES
# ─────────────────────────────────────────────────────────────
@app.get("/")
def root():
    return {"service": "Codiq Engine API", "version": "2.0.0"}


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload one Excel = one question. Returns rows + metadata."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx or .xls files supported.")
    file_id   = str(uuid.uuid4())
    file_path = UPLOAD_DIR / f"{file_id}_{file.filename}"
    contents  = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    try:
        rows, question = read_rows(str(file_path))
    except Exception as e:
        raise HTTPException(422, f"Could not read Excel: {e}")

    return {
        "file_id":  file_id,
        "filename": file.filename,
        "path":     str(file_path),
        "size_kb":  round(len(contents)/1024, 1),
        "total":    len(rows),
        "question": question[:160],
        "all_rows": rows,
        "preview":  rows[:10],
    }


@app.post("/run")
async def run_pipeline(
    background_tasks: BackgroundTasks,
    file_path:     str = Form(...),
    topic:         str = Form(...),
    min_threshold: int = Form(5),
    rows_json:     str = Form(""),      # optional pre-parsed rows
):
    if not os.path.exists(file_path):
        raise HTTPException(404, "File not found. Re-upload.")

    if rows_json:
        input_rows = json.loads(rows_json)
    else:
        input_rows, _ = read_rows(file_path)

    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status":"queued","stage":"Queued","progress":0.0,
        "log":["Job queued…"], "input_path":file_path, "input_rows":input_rows,
        "topic":topic, "all_rows":[], "codebook":[], "stats":{}, "created_at":time.time(),
    }
    runner = real_pipeline if _ollama_alive() else mock_pipeline
    if runner is mock_pipeline:
        jobs[job_id]["log"].append("⚠ Ollama not reachable at localhost:11434 — running demo mock pipeline")
    background_tasks.add_task(runner, job_id, topic, min_threshold)
    return {"job_id": job_id}


@app.get("/status/{job_id}")
def get_status(job_id: str):
    if job_id not in jobs:
        raise HTTPException(404, "Job not found.")
    j = jobs[job_id]
    return {
        "job_id":   job_id,
        "status":   j["status"],
        "stage":    j["stage"],
        "progress": round(j["progress"], 3),
        "log":      j["log"][-40:],
        "stats":    j["stats"],
        "codebook": j["codebook"],
        "all_rows": j["all_rows"],    # ← THE FIX: coded rows now returned
    }


@app.delete("/job/{job_id}")
def cancel_job(job_id: str):
    if job_id not in jobs:
        raise HTTPException(404, "Job not found.")
    jobs[job_id]["status"] = "cancelled"
    return {"cancelled": True}


# ─────────────────────────────────────────────────────────────
#  EXPORT — build coded Excel from workspace state
#  Supports multiple questions (QID column preserved)
# ─────────────────────────────────────────────────────────────
@app.post("/export")
async def export_coded(
    responses: str = Form(...),
    codebook:  str = Form(...),
    question:  str = Form(""),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows  = json.loads(responses)
    codes = json.loads(codebook)

    wb  = Workbook()
    ws1 = wb.active; ws1.title = "Sheet1"
    ws1.cell(1, 1, f"Q: {question}").font = Font(bold=True)

    row = 3
    grouped = {}
    for t in codes:
        grouped.setdefault(t.get("net",900), {"name": t.get("net_name",""), "themes": []})["themes"].append(t)
    for net in sorted(grouped):
        nd = grouped[net]
        ws1.cell(row,1,f"Net {net}: {nd['name']}").font = Font(bold=True)
        ws1.cell(row,3,"Code").font = Font(bold=True)
        row += 1
        for t in nd["themes"]:
            ws1.cell(row,1,f"    {t['theme']}")
            ws1.cell(row,3,t["code"])
            row += 1
        row += 1
    for c, n in {9996:"Other",9997:"Nothing/None/NA",9998:"Don't Know",9999:"Can't Code"}.items():
        ws1.cell(row,1,f"    {n}"); ws1.cell(row,3,c); row += 1
    ws1.column_dimensions["A"].width = 55; ws1.column_dimensions["C"].width = 12

    ws2 = wb.create_sheet("Sheet2")
    headers = ["RESPID","QID","Verbatim","Sentiments","Code_1","Code_2","Code_3","Code_4"]
    hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="1F4E79")
    for ci, h in enumerate(headers, 1):
        c = ws2.cell(1, ci, h); c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal="center")
    fills = {"Positive":PatternFill("solid",fgColor="E2EFDA"),
             "Negative":PatternFill("solid",fgColor="FCE4D6"),
             "Mixed":PatternFill("solid",fgColor="FFF2CC")}
    for ri, r in enumerate(rows, 2):
        ws2.cell(ri,1,r.get("respid",""))
        ws2.cell(ri,2,r.get("qid",""))
        ws2.cell(ri,3,r.get("verbatim",""))
        sc = ws2.cell(ri,4,r.get("sentiment",""))
        if r.get("sentiment") in fills: sc.fill = fills[r["sentiment"]]
        for ci, code in enumerate((r.get("codes") or [])[:4], 5):
            try: ws2.cell(ri, ci, int(code))
            except: pass
    ws2.column_dimensions["A"].width=10; ws2.column_dimensions["B"].width=40
    ws2.column_dimensions["C"].width=60; ws2.column_dimensions["D"].width=12
    for l in ["E","F","G","H"]: ws2.column_dimensions[l].width=9

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=coded_output.xlsx"})