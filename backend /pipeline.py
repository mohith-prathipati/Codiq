"""
╔══════════════════════════════════════════════════════════════╗
║         CODIQ ENGINE v3 — Verbatim Coding Pipeline           ║
║         Universal — works for ANY survey topic               ║
║         Powered by Ollama (Mistral)                          ║
╚══════════════════════════════════════════════════════════════╝

Usage:   python pipeline.py
Requires: ollama pull mistral  |  pip install pandas openpyxl requests tqdm
"""

import json, os, random, re, sys, time
from pathlib import Path
import pandas as pd
import requests
from tqdm import tqdm

# ─────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────
OLLAMA_URL        = "http://localhost:11434/api/generate"
MODEL             = "mistral"
REQUEST_TIMEOUT   = 180
CHUNK_SIZE        = 50
MAX_SAMPLE_CHUNKS = 10
CODE_BATCH        = 15
MAX_CODES         = 4

RESERVED = {9996:"Other", 9997:"Nothing/None/NA", 9998:"Don't Know", 9999:"Can't Code"}
VALID_SENTIMENTS = {"Positive", "Negative", "Mixed"}

NET_RANGES = {
    100:  "Core Product / Service Attributes (Positive)",
    200:  "Functional Benefits",
    300:  "Emotional & Experiential Benefits",
    400:  "Brand & Trust",
    500:  "Design & Aesthetics",
    600:  "Value & Pricing",
    700:  "Naming & Communication Issues",
    800:  "Functional Concerns & Negatives",
    900:  "Emotional & Experiential Concerns",
    1000: "Brand & Trust Concerns",
    1100: "Design & Aesthetics Concerns",
    1200: "Value & Pricing Concerns",
    1300: "Competitive & Market Issues",
    1400: "Target Audience & Relevance Issues",
}
POSITIVE_NETS = [100, 200, 300, 400, 500, 600]
NEGATIVE_NETS = [700, 800, 900, 1000, 1100, 1200, 1300, 1400]

# ─────────────────────────────────────────────────────────────
#  UNIVERSAL RESERVED CODE PATTERNS (topic-agnostic)
# ─────────────────────────────────────────────────────────────
NOTHING_PAT = [
    r"^nothing\.?$", r"^none\.?$", r"^n\/a\.?$", r"^na\.?$", r"^nope\.?$",
    r"^nothing at all\.?$", r"^-$", r"^\.$", r"^nothing special\.?$",
    r"^nothing really\.?$", r"^not sure\.?$",
]
DK_KW = [
    "don't know", "dont know", "not sure", "unsure", "no idea", "idk",
    "i'm not sure", "im not sure", "can't say", "cant say",
    "i am not too sure", "not too sure", "honestly don't know",
    "i don't know what", "i honestly don't know",
]

# ─────────────────────────────────────────────────────────────
#  PROMPTS
# ─────────────────────────────────────────────────────────────
DISCOVER_PROMPT = """You are a senior qualitative research analyst performing verbatim coding.

TASK: Read these survey responses and extract THEMES representing distinct, meaningful ideas.

EXISTING THEMES (do NOT duplicate — only add genuinely NEW ones):
{existing}

RESPONSES (chunk {chunk_num}/{total_chunks}):
{responses}

THRESHOLD: A theme needs at least {threshold} responses expressing this idea to qualify.

RULES:
1. Theme names: short, specific, title-case (3-5 words max)
2. Be GENEROUS — if a concept appears 3+ times with a distinct angle, create it as a theme
3. Each theme must represent a DISTINCT idea — no semantic overlap
4. Do NOT create themes for: Nothing, Don't Know, Other, Can't Code (these are reserved)
5. Return ONLY valid JSON array — the COMPLETE updated list including existing themes

OUTPUT:
[
  {{"theme": "Long Range Capability", "description": "mentions driving range, miles per charge, distance between charges"}},
  {{"theme": "Environmental Benefit", "description": "mentions eco-friendly, green, reducing emissions, better for planet"}}
]
"""

CONSOLIDATE_PROMPT = """You are a senior qualitative research analyst.

TASK: Consolidate these raw themes — merge ONLY near-identical duplicates, keep genuinely distinct ones.

RAW THEMES:
{themes}

RULES:
1. Merge ONLY themes that are essentially the same concept with different wording
2. DO NOT over-merge — keep themes that are meaningfully different even if related
3. Target: 20–35 final themes for a study with 1000+ responses
4. Write a precise INCLUDE/EXCLUDE description for each final theme
5. Return ONLY valid JSON array

OUTPUT:
[
  {{
    "theme": "Long Range Capability",
    "description": "INCLUDE: mentions driving range, miles per charge, battery life between charges. EXCLUDE: general battery concerns without range context."
  }}
]
"""

NET_ASSIGN_PROMPT = """You are a senior qualitative research analyst.

TASK: Assign each theme to the correct NET group and give it a unique 3-digit code.

NET RANGES (universal — applies to any product/service):
{net_ranges}

THEMES:
{themes}

RULES:
1. Assign each theme to its most semantically appropriate NET range
2. Within each NET, number themes sequentially: first = X01, second = X02, etc.
3. No two themes share the same code
4. Codes 100–699 = POSITIVE themes; 700–1499 = NEGATIVE themes
5. Return ONLY valid JSON array

OUTPUT:
[
  {{"code": 201, "net": 200, "net_name": "Functional Benefits", "theme": "Long Range Capability", "description": "INCLUDE: driving range mentions. EXCLUDE: general battery concerns."}}
]
"""

ASSIGN_PROMPT = """You are a senior qualitative research analyst coding survey verbatim responses.

SURVEY TOPIC: {topic}

CODEBOOK — study every entry carefully before coding:
{codebook}

RESERVED CODES — absolute last resort ONLY:
  9996 = Other      → ONLY after checking EVERY code above and NONE fits at all
  9997 = Nothing/NA → ONLY if response is literally "nothing", "none", "n/a"
  9998 = Don't Know → ONLY if literally "don't know", "not sure", "unsure"
  9999 = Can't Code → ONLY if completely gibberish or unreadable

SENTIMENT — ONLY these three values (never "Neutral"):
  Positive = enthusiastic, recommending, praising
  Negative = critical, disappointed, rejecting
  Mixed    = neutral, factual, mentions both pros and cons

CODING STEPS for EVERY response:
1. Read the full response carefully
2. Go through EACH code in the codebook — does this response mention or imply this theme?
3. Think semantically — match the MEANING not just exact words
4. Assign ALL codes that genuinely match (1 to {max_codes} codes)
5. Long responses (>15 words) almost always need 2+ codes — look harder
6. Before using 9996: does the response mention ANY aspect of the topic, product, brand, or experience?
   If yes, there IS a matching code — keep looking.
7. Only use 9996 if the response is truly off-topic or unrelated to the survey subject

RESPONSES:
{responses}

Return ONLY valid JSON array:
[
  {{"respid": "2", "sentiment": "Positive", "codes": [201, 301]}},
  {{"respid": "10", "sentiment": "Negative", "codes": [802]}}
]
"""

RETRY_PROMPT = """You are a senior qualitative research analyst.

SURVEY TOPIC: {topic}

These responses were coded as "Other" (9996). Look MORE carefully — most WILL match a real code.

CODEBOOK:
{codebook}

RESPONSES CODED AS OTHER:
{responses}

For EACH response, ask yourself:
- What is this person talking about? What aspect of {topic} are they referring to?
- Which codebook theme most closely matches what they're saying?
- Even partial matches or implied meanings count.

Only keep 9996 if the response is genuinely unrelated to {topic} or completely uninterpretable.

Return ONLY valid JSON array:
[
  {{"respid": "10", "sentiment": "Negative", "codes": [802]}},
  {{"respid": "180", "sentiment": "Mixed", "codes": [9996]}}
]
"""

KEYWORD_GEN_PROMPT = """You are a qualitative research analyst.

I have built a codebook for a survey about: {topic}

CODEBOOK THEMES:
{themes}

TASK: For each theme, generate a list of keywords and phrases that respondents might use
when expressing that theme — including informal, partial, and indirect phrasings.

Think broadly — include synonyms, abbreviations, colloquial expressions, and related concepts
that a real survey respondent might use.

RULES:
1. Generate 8-15 keywords/phrases per theme
2. Keep each keyword short (1-5 words)
3. Use lowercase
4. Think about how REAL people talk, not just formal definitions
5. Return ONLY valid JSON

OUTPUT:
{{
  "201": ["long range", "driving range", "miles per charge", "range anxiety", "how far", "distance", "battery life", "charge lasts"],
  "301": ["love it", "exciting", "fun to drive", "enjoy", "great experience", "happy with"]
}}
"""

# ─────────────────────────────────────────────────────────────
#  OLLAMA
# ─────────────────────────────────────────────────────────────
def ollama(prompt: str) -> str:
    payload = {
        "model":   MODEL,
        "prompt":  prompt,
        "stream":  False,
        "format":  "json",
        "options": {"temperature": 0.05, "num_predict": 16000},
    }
    for attempt in range(4):
        try:
            r = requests.post(OLLAMA_URL, json=payload, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            return r.json()["response"].strip()
        except Exception as e:
            if attempt == 3:
                raise RuntimeError(f"Ollama failed: {e}")
            time.sleep(3 * (attempt + 1))


def parse_json(text: str):
    text = re.sub(r"```(?:json)?", "", text).strip().rstrip("`").strip()
    for open_c, close_c in [("[", "]"), ("{", "}")]:
        start = text.find(open_c)
        if start == -1: continue
        end = text.rfind(close_c)
        if end == -1: continue
        candidate = text[start:end + 1]
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            last_comma = candidate.rfind(",")
            if last_comma != -1:
                try:
                    return json.loads(candidate[:last_comma] + close_c)
                except Exception:
                    pass
    raise ValueError(f"Cannot parse JSON:\n{text[:400]}")


def safe_int(val):
    try:
        return abs(int(float(str(val).replace("-", "").strip())))
    except Exception:
        return None

# ─────────────────────────────────────────────────────────────
#  STEP 1 — READ INPUT
# ─────────────────────────────────────────────────────────────
def read_input(path: str):
    df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    col_map = {}
    for col in df.columns:
        low = col.lower()
        if any(k in low for k in ["respid","resp_id","response_id"]) and "RESPID" not in col_map:
            col_map["RESPID"] = col
        elif any(k in low for k in ["qid","question_id","q_id"]) and "QID" not in col_map:
            col_map["QID"] = col
        elif any(k in low for k in ["verbatim","response","text","answer"]) and "Verbatim" not in col_map:
            col_map["Verbatim"] = col
    if len(col_map) < 3:
        cols = list(df.columns)
        col_map = {"RESPID": cols[0], "QID": cols[1], "Verbatim": cols[2]}
    df = df.rename(columns={v: k for k, v in col_map.items()})
    df = df[["RESPID","QID","Verbatim"]].copy()
    df["Verbatim"] = df["Verbatim"].fillna("").str.strip()
    df = df[df["Verbatim"] != ""].reset_index(drop=True)
    return df, df["QID"].mode()[0], df["Verbatim"].tolist()

# ─────────────────────────────────────────────────────────────
#  STEP 2 — DISCOVER THEMES
# ─────────────────────────────────────────────────────────────
def discover_themes(verbatims: list, min_threshold: int) -> list:
    all_chunks = [verbatims[i:i + CHUNK_SIZE] for i in range(0, len(verbatims), CHUNK_SIZE)]
    if len(all_chunks) > MAX_SAMPLE_CHUNKS:
        middle  = all_chunks[1:-1]
        sampled = [all_chunks[0]] + random.sample(middle, min(MAX_SAMPLE_CHUNKS - 2, len(middle))) + [all_chunks[-1]]
        print(f"  ↳ {len(verbatims)} responses → sampling {len(sampled)} chunks of {CHUNK_SIZE}")
    else:
        sampled = all_chunks
        print(f"  ↳ {len(verbatims)} responses → {len(sampled)} chunks of {CHUNK_SIZE}")

    existing = []
    for i, chunk in enumerate(tqdm(sampled, desc="  Discovering themes")):
        prompt = DISCOVER_PROMPT.format(
            existing=json.dumps(existing, indent=2) if existing else "[]",
            chunk_num=i + 1,
            total_chunks=len(sampled),
            responses="\n".join(f"- {r}" for r in chunk),
            threshold=min_threshold,
        )
        raw = ollama(prompt)
        try:
            themes = parse_json(raw)
        except Exception as e:
            print(f"\n  ⚠ Chunk {i+1} error: {e}")
            continue
        existing_name_set = {t["theme"].lower() for t in existing}
        for t in themes:
            if isinstance(t, dict) and t.get("theme") and t["theme"].lower() not in existing_name_set:
                existing.append(t)
                existing_name_set.add(t["theme"].lower())
    return existing

# ─────────────────────────────────────────────────────────────
#  STEP 3 — CONSOLIDATE
# ─────────────────────────────────────────────────────────────
def consolidate_themes(raw: list) -> list:
    themes_json = json.dumps(
        [{"theme": t["theme"], "description": t.get("description","")} for t in raw],
        indent=2
    )
    raw_out = ollama(CONSOLIDATE_PROMPT.format(themes=themes_json))
    try:
        result = parse_json(raw_out)
        print(f"  ↳ {len(raw)} raw → {len(result)} consolidated")
        return result
    except Exception as e:
        print(f"  ⚠ Consolidation error: {e}. Using raw.")
        return raw

# ─────────────────────────────────────────────────────────────
#  STEP 4 — ASSIGN NET CODES
# ─────────────────────────────────────────────────────────────
def assign_net_codes(consolidated: list) -> list:
    net_text = "\n".join(f"  {k}–{k+99}: {v}" for k, v in NET_RANGES.items())
    raw = ollama(NET_ASSIGN_PROMPT.format(
        net_ranges=net_text,
        themes=json.dumps(consolidated, indent=2)
    ))
    try:
        codebook = parse_json(raw)
    except Exception as e:
        print(f"  ⚠ NET assignment error: {e}")
        return []
    seen, clean = set(), []
    for t in codebook:
        if not isinstance(t, dict): continue
        code = safe_int(t.get("code"))
        if code and code not in seen and code not in RESERVED:
            t["code"] = code
            t["net"]  = safe_int(t.get("net")) or 800
            seen.add(code)
            clean.append(t)
    clean.sort(key=lambda x: x["code"])
    return clean

# ─────────────────────────────────────────────────────────────
#  STEP 5 — GENERATE DYNAMIC KEYWORD MAP
#  This replaces all hardcoded SUGAR_KW, CALORIE_KW etc.
#  The model generates topic-specific keywords for each theme.
# ─────────────────────────────────────────────────────────────
def generate_keyword_map(codebook: list, topic: str) -> dict:
    """
    Ask the model to generate keyword lists for every theme in the codebook.
    Returns {code: [keyword1, keyword2, ...]} — fully topic-agnostic.
    """
    themes_text = "\n".join(
        f"  {t['code']} = {t['theme']}: {t.get('description','')}"
        for t in codebook
        if t["code"] not in RESERVED
    )
    prompt = KEYWORD_GEN_PROMPT.format(topic=topic, themes=themes_text)
    raw = ollama(prompt)
    try:
        result = parse_json(raw)
        # Normalise: keys should be ints
        kw_map = {}
        for k, v in result.items():
            code = safe_int(k)
            if code and isinstance(v, list):
                kw_map[code] = [str(kw).lower().strip() for kw in v if kw]
        print(f"  ↳ Generated keywords for {len(kw_map)} themes")
        return kw_map
    except Exception as e:
        print(f"  ⚠ Keyword generation error: {e}. Validator injection disabled.")
        return {}

# ─────────────────────────────────────────────────────────────
#  STEP 6 — ASSIGN CODES TO RESPONSES (two-pass)
# ─────────────────────────────────────────────────────────────
def _find_code(codebook: list, *keywords):
    for t in codebook:
        if any(k.lower() in t["theme"].lower() for k in keywords):
            return t["code"]
    return None


def _codebook_text(codebook: list) -> str:
    return "\n".join(
        f"  {t['code']} = {t['theme']}: {t.get('description','')}"
        for t in codebook
    )


def _sanitise(raw_codes: list, valid_codes: set) -> list:
    clean, seen = [], set()
    for c in raw_codes:
        ci = safe_int(c)
        if ci and ci in valid_codes and ci not in seen:
            clean.append(ci); seen.add(ci)
    return clean or [9996]


def _parse_batch(raw: str, valid_codes: set) -> dict:
    results = {}
    try:
        assigned = parse_json(raw)
        if isinstance(assigned, list):
            for item in assigned:
                if not isinstance(item, dict): continue
                rid   = str(item.get("respid",""))
                codes = _sanitise(item.get("codes",[9996]), valid_codes)
                sent  = str(item.get("sentiment","Mixed")).strip().capitalize()
                if sent not in VALID_SENTIMENTS: sent = "Mixed"
                results[rid] = {"sentiment": sent, "codes": codes[:MAX_CODES]}
    except Exception:
        pass
    return results


def assign_codes(df: pd.DataFrame, codebook: list, topic: str) -> pd.DataFrame:
    valid_codes = set(t["code"] for t in codebook) | set(RESERVED.keys())
    cb_text     = _codebook_text(codebook)
    results     = {}

    batches = [df.iloc[i:i+CODE_BATCH] for i in range(0, len(df), CODE_BATCH)]
    for batch in tqdm(batches, desc="  Pass 1: Assigning codes"):
        responses_text = "\n".join(
            f'RESPID {r["RESPID"]}: "{r["Verbatim"]}"' for _, r in batch.iterrows()
        )
        prompt = ASSIGN_PROMPT.format(
            topic=topic,
            codebook=cb_text,
            max_codes=MAX_CODES,
            responses=responses_text,
        )
        results.update(_parse_batch(ollama(prompt), valid_codes))
        for _, row in batch.iterrows():
            if str(row["RESPID"]) not in results:
                results[str(row["RESPID"])] = {"sentiment":"Mixed","codes":[9996]}

    # Pass 2 — retry 9996-only
    retry_rows = [
        {"RESPID": row["RESPID"], "Verbatim": row["Verbatim"]}
        for _, row in df.iterrows()
        if results.get(str(row["RESPID"]),{}).get("codes",[]) == [9996]
        and len(str(row["Verbatim"]).strip()) > 20
    ]
    if retry_rows:
        print(f"\n  → Pass 2: Retrying {len(retry_rows)} responses coded as Other")
        for rb in tqdm([retry_rows[i:i+CODE_BATCH] for i in range(0,len(retry_rows),CODE_BATCH)], desc="  Pass 2"):
            responses_text = "\n".join(
                f'RESPID {r["RESPID"]} [current: 9996]: "{r["Verbatim"]}"' for r in rb
            )
            updated = _parse_batch(
                ollama(RETRY_PROMPT.format(
                    topic=topic,
                    codebook=cb_text,
                    responses=responses_text
                )),
                valid_codes
            )
            for rid, val in updated.items():
                if val["codes"] != [9996]:
                    results[rid] = val

    df = df.copy()
    df["Sentiments"] = "Mixed"
    for i in range(1, MAX_CODES+1): df[f"Code_{i}"] = None
    for idx, row in df.iterrows():
        r = results.get(str(row["RESPID"]), {"sentiment":"Mixed","codes":[9996]})
        df.at[idx,"Sentiments"] = r["sentiment"]
        for i, code in enumerate(r["codes"], 1):
            if i <= MAX_CODES: df.at[idx, f"Code_{i}"] = code
    return df

# ─────────────────────────────────────────────────────────────
#  STEP 7 — VALIDATE & AUTO-CORRECT (fully dynamic)
# ─────────────────────────────────────────────────────────────
def _inject(codes: list, new_code: int) -> list:
    if new_code in codes: return codes
    return ([new_code] + [c for c in codes if c != new_code])[:MAX_CODES]


def validate_and_fix(df: pd.DataFrame, codebook: list, kw_map: dict) -> tuple:
    """
    Dynamic validator — uses AI-generated keyword map instead of
    hardcoded topic-specific keywords. Works for any survey topic.
    """
    valid_codes = set(t["code"] for t in codebook) | set(RESERVED.keys())
    fixes = 0
    df = df.copy()

    for idx, row in df.iterrows():
        verb     = str(row["Verbatim"]).strip()
        verb_low = verb.lower()
        codes    = []
        for col in [f"Code_{i}" for i in range(1, MAX_CODES+1)]:
            v = row[col]
            if pd.notna(v):
                try: codes.append(int(float(v)))
                except: pass
        changed = False

        # 1. Fix invalid sentiment
        if str(row["Sentiments"]).strip() not in VALID_SENTIMENTS:
            df.at[idx,"Sentiments"] = "Mixed"; fixes += 1; changed = True

        # 2. Remove ghost codes
        clean = [c for c in codes if c in valid_codes]
        if len(clean) != len(codes):
            fixes += len(codes)-len(clean); codes = clean or [9996]; changed = True

        # 3. Dedup
        deduped = list(dict.fromkeys(codes))
        if deduped != codes: codes = deduped; fixes += 1; changed = True

        # 4. Force 9997 for nothing/none responses
        if any(re.match(p, verb_low.strip()) for p in NOTHING_PAT):
            if codes != [9997]:
                codes = [9997]; df.at[idx,"Sentiments"] = "Mixed"; fixes += 1; changed = True

        # 5. Force 9998 for don't know responses (short only)
        elif any(k in verb_low for k in DK_KW) and len(verb) < 40:
            if codes != [9998]:
                codes = [9998]; df.at[idx,"Sentiments"] = "Mixed"; fixes += 1; changed = True

        else:
            # 6. Dynamic keyword injection using AI-generated kw_map
            for code, keywords in kw_map.items():
                if code not in valid_codes: continue
                if code in codes: continue
                if any(kw in verb_low for kw in keywords):
                    codes = _inject(codes, code); fixes += 1; changed = True

        if changed:
            for i in range(1, MAX_CODES+1):
                df.at[idx, f"Code_{i}"] = codes[i-1] if i <= len(codes) else None

    return df, fixes

# ─────────────────────────────────────────────────────────────
#  STEP 7b — POST-ASSIGNMENT CODEBOOK DEDUP
# ─────────────────────────────────────────────────────────────
def deduplicate_codebook(codebook: list, coded_df: pd.DataFrame) -> tuple:
    """
    Merge any remaining overlapping themes after assignment.
    Finds pairs of themes in the same NET with very similar names
    and merges the less-used one into the more-used one.
    """
    from collections import defaultdict, Counter

    # Count code usage
    usage = Counter()
    for _, row in coded_df.iterrows():
        for col in [f"Code_{i}" for i in range(1, MAX_CODES+1)]:
            v = row[col]
            if pd.notna(v):
                try: usage[int(float(v))] += 1
                except: pass

    # Find pairs in same NET with word overlap > 0.7
    stops = {"a","an","the","and","or","of","to","in","for","with","by","as","is","are"}
    def sim(a, b):
        aw = set(a.lower().split()) - stops
        bw = set(b.lower().split()) - stops
        if not aw or not bw: return 0.0
        return len(aw & bw) / min(len(aw), len(bw))

    # Group by NET
    by_net = defaultdict(list)
    for t in codebook:
        by_net[t.get("net", 0)].append(t)

    merges = []  # (remove_code, keep_code)
    for net, themes in by_net.items():
        for i, t1 in enumerate(themes):
            for t2 in themes[i+1:]:
                if sim(t1["theme"], t2["theme"]) >= 0.70:
                    # Keep the more-used code
                    if usage.get(t1["code"], 0) >= usage.get(t2["code"], 0):
                        merges.append((t2["code"], t1["code"]))
                    else:
                        merges.append((t1["code"], t2["code"]))

    code_to_theme = {t["code"]: t["theme"] for t in codebook}
    merge_log = []
    for remove_code, keep_code in merges:
        if remove_code not in code_to_theme or keep_code not in code_to_theme:
            continue
        merge_log.append((remove_code, keep_code,
                          code_to_theme[remove_code], code_to_theme[keep_code]))
        for idx, row in coded_df.iterrows():
            codes = [int(float(row[c])) for c in [f"Code_{i}" for i in range(1,MAX_CODES+1)] if pd.notna(row[c])]
            if remove_code in codes:
                new_codes, added = [], False
                for c in codes:
                    if c == remove_code:
                        if keep_code not in new_codes and not added:
                            new_codes.append(keep_code); added = True
                    else:
                        new_codes.append(c)
                new_codes = list(dict.fromkeys(new_codes))[:MAX_CODES]
                for i in range(1, MAX_CODES+1):
                    coded_df.at[idx, f"Code_{i}"] = new_codes[i-1] if i <= len(new_codes) else None
        codebook = [t for t in codebook if t["code"] != remove_code]
        del code_to_theme[remove_code]

    return codebook, coded_df, merge_log

# ─────────────────────────────────────────────────────────────
#  STEP 8 — WRITE OUTPUT EXCEL
# ─────────────────────────────────────────────────────────────
def write_output(question: str, codebook: list, coded_df: pd.DataFrame, output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb        = Workbook()
    HDR_FONT  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    BOLD_FONT = Font(name="Arial", bold=True, size=10)
    NORM_FONT = Font(name="Arial", size=10)
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    NET_FILL  = PatternFill("solid", fgColor="D6E4F0")
    POS_FILL  = PatternFill("solid", fgColor="E2EFDA")
    NEG_FILL  = PatternFill("solid", fgColor="FCE4D6")
    SENT_FILL = {
        "Positive": PatternFill("solid", fgColor="E2EFDA"),
        "Negative": PatternFill("solid", fgColor="FCE4D6"),
        "Mixed":    PatternFill("solid", fgColor="FFF2CC"),
    }

    def cel(ws, r, col, val, font=None, fill=None, align="left"):
        c = ws.cell(row=r, column=col, value=val)
        if font: c.font = font
        if fill: c.fill = fill
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        return c

    ws1 = wb.active; ws1.title = "Sheet1"; row = 1
    ws1.merge_cells(f"A{row}:C{row}")
    cel(ws1, row, 1, f"Q: {question}", BOLD_FONT)
    ws1.row_dimensions[row].height = 30; row += 2

    nets = {}
    for t in codebook:
        nc = t.get("net", 800)
        if nc not in nets:
            nets[nc] = {"name": t.get("net_name", NET_RANGES.get(nc,"")), "themes": []}
        nets[nc]["themes"].append(t)

    def section_hdr(ws, r, label, fill):
        ws.merge_cells(f"A{r}:B{r}"); cel(ws, r, 1, label, BOLD_FONT, fill)
        cel(ws, r, 3, "Code", BOLD_FONT, fill, "center")
        ws.row_dimensions[r].height = 18; return r + 2

    def net_block(ws, r, nc, nd):
        ws.merge_cells(f"A{r}:B{r}"); cel(ws, r, 1, f"Net {nc}: {nd['name']}", BOLD_FONT, NET_FILL)
        cel(ws, r, 3, "", BOLD_FONT, NET_FILL, "center"); ws.row_dimensions[r].height = 16; r += 1
        for t in nd["themes"]:
            ws.merge_cells(f"A{r}:B{r}"); cel(ws, r, 1, f"    {t['theme']}", NORM_FONT)
            cel(ws, r, 3, t["code"], NORM_FONT, None, "center"); ws.row_dimensions[r].height = 15; r += 1
        return r + 1

    row = section_hdr(ws1, row, "Net: Positives / Recommend for", POS_FILL)
    for nc in POSITIVE_NETS:
        if nc in nets: row = net_block(ws1, row, nc, nets[nc])
    row = section_hdr(ws1, row, "Net: Negatives / Suggestions to Improve", NEG_FILL)
    for nc in NEGATIVE_NETS:
        if nc in nets: row = net_block(ws1, row, nc, nets[nc])
    for nc in sorted(n for n in nets if n not in POSITIVE_NETS and n not in NEGATIVE_NETS):
        row = net_block(ws1, row, nc, nets[nc])
    row += 1
    for code, name in RESERVED.items():
        ws1.merge_cells(f"A{row}:B{row}"); cel(ws1, row, 1, f"    {name}", NORM_FONT)
        cel(ws1, row, 3, code, NORM_FONT, None, "center"); ws1.row_dimensions[row].height = 15; row += 1

    ws1.column_dimensions["A"].width = 58
    ws1.column_dimensions["B"].width = 5
    ws1.column_dimensions["C"].width = 12

    ws2 = wb.create_sheet("Sheet2")
    for ci, h in enumerate(["RESPID","QID","Verbatim","Sentiments","Code_1","Code_2","Code_3","Code_4"], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 18

    for ri, (_, r) in enumerate(coded_df.iterrows(), 2):
        ws2.cell(row=ri, column=1, value=r["RESPID"])
        ws2.cell(row=ri, column=2, value=r["QID"])
        ws2.cell(row=ri, column=3, value=r["Verbatim"])
        sc = ws2.cell(row=ri, column=4, value=r["Sentiments"])
        if r["Sentiments"] in SENT_FILL: sc.fill = SENT_FILL[r["Sentiments"]]
        for ci, col in enumerate(["Code_1","Code_2","Code_3","Code_4"], 5):
            val = r.get(col)
            if pd.notna(val) and val is not None:
                try: ws2.cell(row=ri, column=ci, value=int(float(val)))
                except: pass

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 80
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 12
    for l in ["E","F","G","H"]: ws2.column_dimensions[l].width = 10

    wb.save(output_path)
    print(f"\n  ✅ Output saved → {output_path}")

# ─────────────────────────────────────────────────────────────
#  JSON I/O
# ─────────────────────────────────────────────────────────────
def save_codebook(cb, path):
    with open(path,"w",encoding="utf-8") as f: json.dump(cb, f, indent=2, ensure_ascii=False)
    print(f"  ✅ Codebook saved → {path}")

def load_codebook(path):
    with open(path,"r",encoding="utf-8") as f: return json.load(f)

def save_kw_map(kw_map, path):
    with open(path,"w",encoding="utf-8") as f:
        json.dump({str(k): v for k,v in kw_map.items()}, f, indent=2)
    print(f"  ✅ Keyword map saved → {path}")

def load_kw_map(path):
    with open(path,"r",encoding="utf-8") as f:
        return {int(k): v for k,v in json.load(f).items()}

# ─────────────────────────────────────────────────────────────
#  STATS
# ─────────────────────────────────────────────────────────────
def print_stats(coded_df: pd.DataFrame, codebook: list):
    valid_codes = set(t["code"] for t in codebook) | set(RESERVED.keys())
    ghost = sum(
        1 for _, row in coded_df.iterrows()
        for col in [f"Code_{i}" for i in range(1,MAX_CODES+1)]
        if pd.notna(row[col]) and safe_int(row[col]) not in valid_codes
    )
    reserved_only = sum(
        1 for _, row in coded_df.iterrows()
        if any(pd.notna(row[f"Code_{i}"]) for i in range(1,MAX_CODES+1))
        and all(
            not pd.notna(row[f"Code_{i}"]) or int(float(row[f"Code_{i}"])) in RESERVED
            for i in range(1,MAX_CODES+1)
        )
    )
    print(f"\n── Run Summary ────────────────────────────────────────")
    print(f"  Responses coded      : {len(coded_df)}")
    for s, n in coded_df["Sentiments"].value_counts().items():
        print(f"  {s:12s}       : {n:4d}  ({n/len(coded_df)*100:.1f}%)")
    print(f"\n  Ghost codes          : {ghost}")
    print(f"  Invalid sentiments   : {coded_df[~coded_df['Sentiments'].isin(VALID_SENTIMENTS)].shape[0]}")
    print(f"  Reserved-only coded  : {reserved_only}")
    dist = {}
    for _, row in coded_df.iterrows():
        n = sum(1 for c in [f"Code_{i}" for i in range(1,MAX_CODES+1)] if pd.notna(row[c]))
        dist[n] = dist.get(n,0)+1
    print("\n  Codes per response:")
    for n, cnt in sorted(dist.items()):
        print(f"    {n} codes : {cnt:4d}  ({cnt/len(coded_df)*100:.1f}%)")

# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────
def main():
    print("\n" + "═"*62)
    print("   CODIQ ENGINE v3 — Universal Verbatim Coding Pipeline")
    print("   Powered by Ollama / Mistral")
    print("═"*62 + "\n")

    input_path = input("📂 Input Excel file path: ").strip().strip('"')
    if not os.path.exists(input_path):
        print(f"❌ File not found: {input_path}"); sys.exit(1)

    topic = input("🔍 Survey topic (e.g. 'Electric Vehicles', 'Coca-Cola Zero Sugar'): ").strip()
    if not topic:
        topic = "the product being surveyed"

    threshold_str = input("🔢 Min responses to create a theme [default: 5]: ").strip()
    min_threshold = int(threshold_str) if threshold_str.isdigit() else 5

    existing_cb = input("📋 Existing codebook JSON? (leave blank to build fresh): ").strip().strip('"')

    default_out = str(Path(input_path).parent / f"{Path(input_path).stem}_coded_v3.xlsx")
    output_path = input(f"💾 Output path [default: {default_out}]: ").strip().strip('"') or default_out

    print("\n── Reading input ──────────────────────────────────────")
    df, question, verbatims = read_input(input_path)
    print(f"  ✅ {len(df)} responses loaded")
    print(f"  📌 Question: {question[:90]}...")

    if existing_cb and os.path.exists(existing_cb):
        print("\n── Loading existing codebook ──────────────────────────")
        codebook = load_codebook(existing_cb)
        print(f"  ✅ {len(codebook)} themes loaded")

        # Try loading saved keyword map
        kw_path = existing_cb.replace("_codebook_", "_kwmap_").replace(".json","_kwmap.json")
        if os.path.exists(kw_path):
            kw_map = load_kw_map(kw_path)
            print(f"  ✅ Keyword map loaded ({len(kw_map)} themes)")
        else:
            print("\n── Generating keyword map ─────────────────────────────")
            kw_map = generate_keyword_map(codebook, topic)
    else:
        print("\n── Stage 1: Discovering themes ────────────────────────")
        raw_themes = discover_themes(verbatims, min_threshold)
        print(f"  ✅ {len(raw_themes)} raw themes discovered")

        print("\n── Stage 2: Consolidating themes ──────────────────────")
        consolidated = consolidate_themes(raw_themes)

        print("\n── Stage 3: Assigning NET codes ───────────────────────")
        codebook = assign_net_codes(consolidated)
        print(f"  ✅ {len(codebook)} final themes across {len(set(t['net'] for t in codebook))} NETs")

        cb_path = str(Path(output_path).parent / f"{Path(input_path).stem}_codebook_v3.json")
        save_codebook(codebook, cb_path)

        print("\n── Stage 4: Generating keyword map ────────────────────")
        kw_map = generate_keyword_map(codebook, topic)
        kw_path = str(Path(output_path).parent / f"{Path(input_path).stem}_kwmap_v3.json")
        save_kw_map(kw_map, kw_path)

    print("\n── Codebook ───────────────────────────────────────────")
    cur_net = None
    for t in codebook:
        if t["net"] != cur_net:
            cur_net = t["net"]
            print(f"\n  NET {t['net']}: {t.get('net_name','')}")
        print(f"    {t['code']}  {t['theme']}")
    print()

    print("\n── Assigning codes ────────────────────────────────────")
    print(f"  ⚙  Batch: {CODE_BATCH} | Responses: {len(df)}")
    coded_df = assign_codes(df, codebook, topic)

    print("\n── Validating & auto-correcting ───────────────────────")
    coded_df, fixes = validate_and_fix(coded_df, codebook, kw_map)
    print(f"  ✅ {fixes} issues auto-corrected")

    print("\n── Merging overlapping themes ─────────────────────────")
    codebook, coded_df, merges = deduplicate_codebook(codebook, coded_df)
    if merges:
        for rem, keep, rn, kn in merges:
            print(f"  ↳ Merged {rem} ({rn}) → {keep} ({kn})")
    else:
        print("  ↳ No overlapping themes found")

    print("\n── Writing output ─────────────────────────────────────")
    write_output(question, codebook, coded_df, output_path)

    print_stats(coded_df, codebook)
    print("\n  🎉 Done!\n")

if __name__ == "__main__":
    main()